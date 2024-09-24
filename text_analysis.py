import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from nltk.tokenize import sent_tokenize, word_tokenize, RegexpTokenizer
import pandas as pd

# Reading Input File
input_file_path="Input.xlsx"
workbook = load_workbook(input_file_path)
input_sheet = workbook['Sheet1']
input_total_rows=input_sheet.max_row
input_total_cols=input_sheet.max_column

# Reading Output File
output_file_path="Output Data Structure.xlsx"
workbook = load_workbook(output_file_path)
output_sheet = workbook['Sheet1']
output_total_rows=output_sheet.max_row
output_total_cols=output_sheet.max_column

# Feeding Excel Data in List
excel_to_list = []
for i in range(0, input_total_rows):
    each_row = []
    for j in range(0, input_total_cols):
        each_row.append(input_sheet.cell(row = i+1, column = j+1).value)
    excel_to_list.append(each_row)

# Creating a List to hold final data to be fed to Output File
list_to_excel = excel_to_list.copy()

# Utility Function to convert List to String
'''
def list_to_string(lst):
    list_to_string = ""
    length_lst = len(lst)
    for i in range(length_lst):
        if i != length_lst-1:
            list_to_string = list_to_string + " " + lst[i]
'''

# Function to Scrape Text using given URL
def url_to_text(url):
    text_to_analyze = ""
    response = requests.get(url)
    html_content = response.content

    soup = BeautifulSoup(html_content, "html.parser")

    # Two div classes have been identified
    division1 = soup.find("div", {"class": "td-post-content tagdiv-type"})
    division2 = soup.find("div", {"class": "tdb-block-inner td-fix-index"})

    division = None

    # If the desired content in either of the div classes
    if division1:
        division = division1
    else:
        division = division2

    # Collecting all the paragraphs inside class
    if division:
        all_paragraphs = division.find_all("p")
        paragraph_list =[]
        for paragraph in all_paragraphs:
            paragraph_list.append(paragraph.text)
            
    # Converting from List to String
    list_to_string = ""
    length_list = len(paragraph_list)
    for i in range(length_list):
        if i != length_list-1:
            list_to_string = list_to_string + " " + paragraph_list[i]
    return list_to_string

# Function to remove Stop Words
def stopword_remover(text_to_analyze):
    path = "StopWords/"
    all_files = os.listdir(path)
    stop_words=[]
    for filename in all_files:
        with open(path+filename, "r") as f:
            stop_words.extend(f.readlines())

    text_without_stop_words = [word for word in text_to_analyze if word not in stop_words]

    list_to_string = ""
    length_list = len(text_without_stop_words)
    for i in range(length_list):
        if i != length_list-1:
            list_to_string = list_to_string + " " + text_without_stop_words[i]
    return list_to_string

# Function to get sentences
def sent_tokenizer(text_to_analyze):
    return sent_tokenize(text_to_analyze)

# Function to get words
def word_tokenizer(text_to_analyze):
    return word_tokenize(text_to_analyze)

# Function to remove punctuation
def punctuation_remover(text_to_analyze):
    without_punctuation_list = RegexpTokenizer(r'\w+').tokenize(text_to_analyze)

    list_to_string = ""
    length_list = len(without_punctuation_list)
    for i in range(length_list):
        if i != length_list-1:
            list_to_string = list_to_string + " " + without_punctuation_list[i]
    return list_to_string

# Function to count number of characters
def text_length(text_to_analyze):
    for word in text_to_analyze:
        count = 0
        for char in word:
            count += 1
    return count

# Function to count number of pronouns
def personal_pronoun_counter(text_to_analyze):
    personal_pronoun_counter = 0
    for word in text_to_analyze:
        if word != "US" and word.lower in ["i", "me", "my", "mine", "we", "us", "our", "ours", "you", "your", "yours", "he", "his", "him", "she", "her", "hers", "they", "them", "their"]:
            personal_pronoun_counter += 1
    return personal_pronoun_counter

# Function to count number of syllables in each word, outputting a List
def syllable_counter(text_to_analyze):
    syllable_counter_list = []
    for word in text_to_analyze:
        syllable_counter = 0
        if word[-2:] != "es" and word[-2:] != "ed":
            for char in word:
                if char in 'aeiou':
                    syllable_counter += 1
        syllable_counter_list.append(syllable_counter)
    return syllable_counter_list

# Function to convert List of syllable count to string to be fed to Excel
def syllable_counter_string(syllable_counter_list):
    list_to_string = ""
    length_list = len(syllable_counter_list)
    for i in range(length_list):
        if i != length_list-1:
            list_to_string = list_to_string + " " + str(syllable_counter_list[i])
    return list_to_string

# Function to count number of complex words
def complex_word_counter(syllable_counter_list):
    complex_word_count = 0
    for i in syllable_counter_list:
        if i == 2:
            complex_word_count += 1
    return complex_word_count

# Function to calculate Positive Score
def positive_score(text_to_analyze):
    path = "MasterDictionary/positive-words.txt"
    positive_words=[]
    count = 0
    with open(path, "r") as f:
        positive_words.extend(f.readlines())
    for word in text_to_analyze:
        if word+"\n" in positive_words:
            count += 1
    return count

# Function to calculate Negative Score
def negative_score(text_to_analyze):
    path = "MasterDictionary/negative-words.txt"
    negative_words=[]
    count = 0
    with open(path, "r") as f:
        negative_words.extend(f.readlines())
    for word in text_to_analyze:
        if word+"\n" in negative_words:
            count += 1
    return count

# Saving Header Row as first item in nested List, made to contain the final data
first_row = []
for data in output_sheet["1"]:
    first_row.append(data.value)
list_to_excel[0] = first_row

for i in range(1, len(excel_to_list)):
    text_to_analyze=url_to_text(excel_to_list[i][1])
    print("----------------- URL -> ", excel_to_list[i][1], " -----------------------")

    # If some text is scraped
    if text_to_analyze:
        text_to_analyze = stopword_remover(word_tokenizer(text_to_analyze))
        num_of_sentences = len(sent_tokenize(text_to_analyze))
        text_to_analyze = punctuation_remover(text_to_analyze)
        text_to_analyze = word_tokenizer(text_to_analyze)
        total_words_after_cleaning = len(text_to_analyze)
        average_word_length = text_length(text_to_analyze) / total_words_after_cleaning
        personal_pronoun_count = personal_pronoun_counter(text_to_analyze)
        syllable_count_list = syllable_counter(text_to_analyze)
        complex_word_count = complex_word_counter(syllable_count_list)
        positive_sc = positive_score(text_to_analyze)
        negative_sc = negative_score(text_to_analyze)
        polarity_score = (positive_sc - negative_sc) / ((positive_sc + negative_sc) + 0.000001)
        subjectivity_score = (positive_sc - negative_sc) / ((total_words_after_cleaning) + 0.000001)
        average_sentence_length = total_words_after_cleaning / num_of_sentences
        percentage_of_complex_words = complex_word_count / total_words_after_cleaning
        fog_index = 0.4 * (average_sentence_length + percentage_of_complex_words)
        average_number_of_words_per_sentence = total_words_after_cleaning / num_of_sentences

        #Saving metrics as successive column values for each row
        list_to_excel[i].append(positive_sc)
        list_to_excel[i].append(negative_sc)
        list_to_excel[i].append(polarity_score)
        list_to_excel[i].append(subjectivity_score)
        list_to_excel[i].append(average_sentence_length)
        list_to_excel[i].append(percentage_of_complex_words)
        list_to_excel[i].append(fog_index)
        list_to_excel[i].append(average_number_of_words_per_sentence)
        list_to_excel[i].append(complex_word_count)
        list_to_excel[i].append(total_words_after_cleaning)
        list_to_excel[i].append(syllable_counter_string(syllable_count_list))
        list_to_excel[i].append(personal_pronoun_count)
        list_to_excel[i].append(average_word_length)

    # If no text could be scraped, save that row with default value i.e., 0
    else:
        for _ in range(13):
            list_to_excel[i].append(0)
    print(list_to_excel)

# Exporting final List to Excel (CSV).
df = pd.DataFrame(list_to_excel)
df.to_csv("Output.csv", index=False, header=False)
